import openpyxl
from openpyxl.utils import get_column_letter
import json
import matplotlib.pyplot as plt
import base64
from io import BytesIO
from typing import List, Dict, Any, Optional, Iterator


class ExcelProcessor:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None

    def load_workbook(self):
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)

    def _get_merged_cell_value(self, sheet, row: int, col: int) -> Any:
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and \
               merged_range.min_col <= col <= merged_range.max_col:
                return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        return sheet.cell(row=row, column=col).value

    def get_sheet_names(self) -> List[str]:
        if self.workbook is None:
            self.load_workbook()
        return self.workbook.sheetnames

    def extract_range_data(
        self,
        sheet_name: str,
        start_row: int = 1,
        end_row: Optional[int] = None,
        start_col: int = 1,
        end_col: Optional[int] = None,
        has_header: bool = True
    ) -> Dict[str, Any]:
        if self.workbook is None:
            self.load_workbook()
        
        sheet = self.workbook[sheet_name]
        
        if end_row is None:
            end_row = sheet.max_row
        if end_col is None:
            end_col = sheet.max_column
        
        headers = []
        data = []
        
        for col in range(start_col, end_col + 1):
            col_letter = get_column_letter(col)
            headers.append(self._get_merged_cell_value(sheet, start_row, col) if has_header else f"Column_{col_letter}")
        
        start_data_row = start_row + 1 if has_header else start_row
        
        for row in range(start_data_row, end_row + 1):
            row_data = {}
            for col_idx, col in enumerate(range(start_col, end_col + 1)):
                cell_value = self._get_merged_cell_value(sheet, row, col)
                header = headers[col_idx] if col_idx < len(headers) else f"Column_{get_column_letter(col)}"
                row_data[header] = cell_value
            data.append(row_data)
        
        return {
            "sheet_name": sheet_name,
            "headers": headers,
            "data": data,
            "row_count": len(data),
            "column_count": len(headers)
        }

    def to_json(self, data: Dict[str, Any], indent: int = 2) -> str:
        return json.dumps(data, ensure_ascii=False, indent=indent, default=str)

    def generate_chart(
        self,
        data: Dict[str, Any],
        chart_type: str = "bar",
        x_column: Optional[str] = None,
        y_columns: Optional[List[str]] = None,
        title: str = "Chart",
        figsize: tuple = (10, 6)
    ) -> str:
        headers = data["headers"]
        rows = data["data"]
        
        if not rows:
            raise ValueError("No data available for chart generation")
        
        if x_column is None:
            x_column = headers[0]
        
        if y_columns is None:
            if len(headers) > 1:
                y_columns = headers[1:3]
            else:
                y_columns = headers
        
        x_data = [row[x_column] for row in rows]
        
        fig, ax = plt.subplots(figsize=figsize)
        
        if chart_type == "bar":
            for y_col in y_columns:
                y_data = [row[y_col] for row in rows]
                ax.bar(x_data, y_data, label=y_col)
        elif chart_type == "line":
            for y_col in y_columns:
                y_data = [row[y_col] for row in rows]
                ax.plot(x_data, y_data, marker='o', label=y_col)
        else:
            raise ValueError(f"Unsupported chart type: {chart_type}. Use 'bar' or 'line'")
        
        ax.set_xlabel(x_column)
        ax.set_ylabel("Value")
        ax.set_title(title)
        ax.legend()
        plt.xticks(rotation=45)
        plt.tight_layout()
        
        buffer = BytesIO()
        plt.savefig(buffer, format='png')
        buffer.seek(0)
        image_base64 = base64.b64encode(buffer.read()).decode('utf-8')
        plt.close()
        
        return image_base64

    def generate_bar_chart(
        self,
        data: Dict[str, Any],
        x_column: Optional[str] = None,
        y_columns: Optional[List[str]] = None,
        title: str = "Bar Chart"
    ) -> str:
        return self.generate_chart(data, "bar", x_column, y_columns, title)

    def generate_line_chart(
        self,
        data: Dict[str, Any],
        x_column: Optional[str] = None,
        y_columns: Optional[List[str]] = None,
        title: str = "Line Chart"
    ) -> str:
        return self.generate_chart(data, "line", x_column, y_columns, title)

    def extract_multiple_sheets(
        self,
        sheet_names: Optional[List[str]] = None,
        start_row: int = 1,
        end_row: Optional[int] = None,
        start_col: int = 1,
        end_col: Optional[int] = None,
        has_header: bool = True
    ) -> Dict[str, Any]:
        if self.workbook is None:
            self.load_workbook()
        
        if sheet_names is None:
            sheet_names = self.get_sheet_names()
        
        results = {}
        for sheet_name in sheet_names:
            try:
                results[sheet_name] = self.extract_range_data(
                    sheet_name, start_row, end_row, start_col, end_col, has_header
                )
            except Exception as e:
                results[sheet_name] = {"error": str(e)}
        
        return {
            "file_path": self.file_path,
            "sheet_count": len(sheet_names),
            "sheets": results
        }

    def to_html_table(
        self,
        data: Dict[str, Any],
        include_style: bool = True,
        table_class: str = "excel-table"
    ) -> str:
        headers = data["headers"]
        rows = data["data"]
        
        style = ""
        if include_style:
            style = """
        <style>
            .excel-table {
                border-collapse: collapse;
                width: 100%;
                font-family: Arial, sans-serif;
                font-size: 14px;
            }
            .excel-table th {
                background-color: #4472C4;
                color: white;
                padding: 12px 15px;
                text-align: left;
                border: 1px solid #305496;
                font-weight: bold;
            }
            .excel-table td {
                padding: 8px 15px;
                border: 1px solid #D4D4D4;
            }
            .excel-table tr:nth-child(even) {
                background-color: #F2F2F2;
            }
            .excel-table tr:hover {
                background-color: #E7F3FF;
            }
        </style>
        """
        
        html = f"{style}\n<table class='{table_class}'>\n"
        
        html += "  <thead>\n    <tr>\n"
        for header in headers:
            html += f"      <th>{header}</th>\n"
        html += "    </tr>\n  </thead>\n"
        
        html += "  <tbody>\n"
        for row in rows:
            html += "    <tr>\n"
            for header in headers:
                value = row.get(header, "")
                html += f"      <td>{value}</td>\n"
            html += "    </tr>\n"
        html += "  </tbody>\n</table>"
        
        return html

    def read_paginated(
        self,
        sheet_name: str,
        page_size: int = 10000,
        start_row: int = 1,
        end_row: Optional[int] = None,
        start_col: int = 1,
        end_col: Optional[int] = None,
        has_header: bool = True
    ) -> Iterator[Dict[str, Any]]:
        if self.workbook is None:
            self.load_workbook()
        
        sheet = self.workbook[sheet_name]
        
        if end_row is None:
            end_row = sheet.max_row
        if end_col is None:
            end_col = sheet.max_column
        
        headers = []
        if has_header:
            for col in range(start_col, end_col + 1):
                headers.append(self._get_merged_cell_value(sheet, start_row, col))
            first_data_row = start_row + 1
        else:
            for col in range(start_col, end_col + 1):
                col_letter = get_column_letter(col)
                headers.append(f"Column_{col_letter}")
            first_data_row = start_row
        
        current_row = first_data_row
        page_number = 1
        
        while current_row <= end_row:
            page_end_row = min(current_row + page_size - 1, end_row)
            
            page_data = []
            for row in range(current_row, page_end_row + 1):
                row_data = {}
                for col_idx, col in enumerate(range(start_col, end_col + 1)):
                    cell_value = self._get_merged_cell_value(sheet, row, col)
                    header = headers[col_idx] if col_idx < len(headers) else f"Column_{get_column_letter(col)}"
                    row_data[header] = cell_value
                page_data.append(row_data)
            
            yield {
                "sheet_name": sheet_name,
                "page_number": page_number,
                "page_size": page_size,
                "total_pages": (end_row - first_data_row + page_size) // page_size,
                "start_row": current_row,
                "end_row": page_end_row,
                "headers": headers,
                "data": page_data,
                "row_count": len(page_data)
            }
            
            current_row = page_end_row + 1
            page_number += 1


def process_excel(
    file_path: str,
    sheet_name: str,
    start_row: int = 1,
    end_row: Optional[int] = None,
    start_col: int = 1,
    end_col: Optional[int] = None,
    has_header: bool = True,
    chart_type: Optional[str] = None,
    x_column: Optional[str] = None,
    y_columns: Optional[List[str]] = None,
    chart_title: Optional[str] = None
) -> Dict[str, Any]:
    processor = ExcelProcessor(file_path)
    
    data = processor.extract_range_data(
        sheet_name, start_row, end_row, start_col, end_col, has_header)
    
    result = {
        "data": data,
        "json": processor.to_json(data)
    }
    
    if chart_type:
        if chart_title is None:
            chart_title = f"{chart_type.capitalize()} Chart"
        chart_base64 = processor.generate_chart(
            data, chart_type, x_column, y_columns, chart_title)
        result["chart_base64"] = chart_base64
        result["chart_type"] = chart_type
    
    return result


if __name__ == "__main__":
    pass
