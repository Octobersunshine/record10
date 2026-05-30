import openpyxl
from openpyxl.styles import Font, PatternFill
import random


def create_sample_excel(file_path: str = "sample_data.xlsx"):
    wb = openpyxl.Workbook()
    
    ws1 = wb.active
    ws1.title = "销售数据"
    
    headers = ["月份", "产品A销量", "产品B销量", "产品C销量", "总销售额"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    months = ["1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月"]
    
    for row, month in enumerate(months, 2):
        ws1.cell(row=row, column=1, value=month)
        ws1.cell(row=row, column=2, value=random.randint(100, 500))
        ws1.cell(row=row, column=3, value=random.randint(80, 450))
        ws1.cell(row=row, column=4, value=random.randint(120, 600))
        ws1.cell(row=row, column=5, value=f"=SUM(B{row}:D{row})*100")
    
    ws2 = wb.create_sheet("员工信息")
    
    headers2 = ["员工ID", "姓名", "部门", "年龄", "薪资"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    employees = [
        (1, "张三", "技术部", 28, 15000),
        (2, "李四", "市场部", 32, 18000),
        (3, "王五", "财务部", 26, 12000),
        (4, "赵六", "技术部", 35, 22000),
        (5, "钱七", "人事部", 29, 13000),
        (6, "孙八", "技术部", 24, 10000),
    ]
    
    for row, emp in enumerate(employees, 2):
        for col, value in enumerate(emp, 1):
            ws2.cell(row=row, column=col, value=value)
    
    wb.save(file_path)
    print(f"示例Excel文件已创建: {file_path}")
    print(f"包含工作表: {wb.sheetnames}")


if __name__ == "__main__":
    create_sample_excel()
