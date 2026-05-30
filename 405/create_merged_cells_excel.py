import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def create_merged_cells_excel(file_path: str = "merged_cells_test.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合并单元格测试"
    
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1, value="2024年度销售报告")
    cell.font = Font(bold=True, size=16)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:A3')
    ws.cell(row=2, column=1, value="季度")
    
    ws.merge_cells('B2:D2')
    ws.cell(row=2, column=2, value="产品销量")
    
    ws.cell(row=3, column=2, value="产品A")
    ws.cell(row=3, column=3, value="产品B")
    ws.cell(row=3, column=4, value="产品C")
    
    ws.merge_cells('A4:A6')
    ws.cell(row=4, column=1, value="Q1")
    ws.cell(row=4, column=2, value=100)
    ws.cell(row=4, column=3, value=150)
    ws.cell(row=4, column=4, value=200)
    ws.cell(row=5, column=2, value=120)
    ws.cell(row=5, column=3, value=180)
    ws.cell(row=5, column=4, value=220)
    ws.cell(row=6, column=2, value=140)
    ws.cell(row=6, column=3, value=160)
    ws.cell(row=6, column=4, value=240)
    
    ws.merge_cells('A7:A9')
    ws.cell(row=7, column=1, value="Q2")
    ws.cell(row=7, column=2, value=160)
    ws.cell(row=7, column=3, value=190)
    ws.cell(row=7, column=4, value=260)
    ws.cell(row=8, column=2, value=180)
    ws.cell(row=8, column=3, value=210)
    ws.cell(row=8, column=4, value=280)
    ws.cell(row=9, column=2, value=200)
    ws.cell(row=9, column=3, value=230)
    ws.cell(row=9, column=4, value=300)
    
    ws.merge_cells('A10:C10')
    cell = ws.cell(row=10, column=1, value="备注: 以上数据仅供参考")
    cell.font = Font(italic=True, color="666666")
    cell.alignment = Alignment(horizontal='left')
    
    ws.merge_cells('D10:D11')
    ws.cell(row=10, column=4, value="总计")
    ws.cell(row=11, column=1, value="合计")
    ws.cell(row=11, column=2, value=1000)
    ws.cell(row=11, column=3, value=1120)
    
    for row in ws.iter_rows(min_row=1, max_row=11, min_col=1, max_col=4):
        for cell in row:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
    
    wb.save(file_path)
    print(f"包含合并单元格的测试Excel已创建: {file_path}")
    print(f"合并单元格区域: {list(ws.merged_cells.ranges)}")


if __name__ == "__main__":
    create_merged_cells_excel()
