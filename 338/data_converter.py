import csv
import json
import io
import sys
import os
import time
import uuid
import codecs
import xml.etree.ElementTree as ET
from typing import List, Dict, Any, Union, Optional, Callable

BOM = '\ufeff'
LARGE_FILE_THRESHOLD = 100 * 1024 * 1024
CHUNK_SIZE = 10000

_tasks: Dict[str, Dict[str, Any]] = {}


def register_task(task_id: Optional[str] = None) -> str:
    if task_id is None:
        task_id = str(uuid.uuid4())
    _tasks[task_id] = {
        'task_id': task_id,
        'status': 'pending',
        'progress': 0,
        'message': 'Task registered',
        'start_time': None,
        'end_time': None,
        'current_chunk': 0,
        'total_chunks': 0,
        'error': None,
        'result': None,
    }
    return task_id


def update_progress(task_id: str, progress: int, message: str = '',
                    current_chunk: int = None, total_chunks: int = None) -> None:
    if task_id not in _tasks:
        return
    task = _tasks[task_id]
    task['progress'] = max(0, min(100, progress))
    if message:
        task['message'] = message
    if current_chunk is not None:
        task['current_chunk'] = current_chunk
    if total_chunks is not None:
        task['total_chunks'] = total_chunks


def mark_task_started(task_id: str) -> None:
    if task_id not in _tasks:
        return
    _tasks[task_id]['status'] = 'running'
    _tasks[task_id]['start_time'] = time.time()


def mark_task_completed(task_id: str, result: Any = None, message: str = 'Task completed') -> None:
    if task_id not in _tasks:
        return
    _tasks[task_id]['status'] = 'completed'
    _tasks[task_id]['progress'] = 100
    _tasks[task_id]['end_time'] = time.time()
    _tasks[task_id]['message'] = message
    _tasks[task_id]['result'] = result


def mark_task_failed(task_id: str, error: str) -> None:
    if task_id not in _tasks:
        return
    _tasks[task_id]['status'] = 'failed'
    _tasks[task_id]['end_time'] = time.time()
    _tasks[task_id]['error'] = error
    _tasks[task_id]['message'] = f'Failed: {error}'


def get_task_status(task_id: str) -> Optional[Dict[str, Any]]:
    return _tasks.get(task_id)


def list_tasks() -> List[Dict[str, Any]]:
    return list(_tasks.values())


def strip_bom(data: str) -> str:
    if data and data[0] == BOM:
        return data[1:]
    return data


def read_file(path: str) -> str:
    with open(path, 'r', encoding='utf-8-sig') as f:
        return f.read()


def write_file(path: str, data: str) -> None:
    with open(path, 'w', encoding='utf-8') as f:
        f.write(data)


def detect_csv_delimiter(data: str) -> str:
    data = strip_bom(data)
    delimiters = [',', '\t', ';']
    first_line = data.splitlines()[0] if data.splitlines() else ''
    max_count = -1
    best_delimiter = ','

    for d in delimiters:
        count = first_line.count(d)
        if count > max_count:
            max_count = count
            best_delimiter = d

    return best_delimiter


def csv_to_json(csv_data: str, delimiter: str = None) -> str:
    csv_data = strip_bom(csv_data)

    if delimiter is None:
        delimiter = detect_csv_delimiter(csv_data)

    reader = csv.DictReader(io.StringIO(csv_data), delimiter=delimiter)
    rows = list(reader)

    return json.dumps(rows, ensure_ascii=False, indent=2)


def json_to_csv(json_data: str, delimiter: str = ',') -> str:
    data = json.loads(json_data)

    if not isinstance(data, list):
        data = [data]

    if not data:
        return ''

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys(), delimiter=delimiter, lineterminator='\n')
    writer.writeheader()
    writer.writerows(data)

    return output.getvalue()


def json_to_xml(json_data: str, root_tag: str = 'root', item_tag: str = 'item') -> str:
    data = json.loads(json_data)

    root = ET.Element(root_tag)

    if isinstance(data, list):
        for item in data:
            item_elem = ET.SubElement(root, item_tag)
            _dict_to_xml(item, item_elem)
    else:
        _dict_to_xml(data, root)

    ET.indent(root, space='  ')
    return ET.tostring(root, encoding='unicode', xml_declaration=True)


def _dict_to_xml(data: Dict[str, Any], parent: ET.Element) -> None:
    for key, value in data.items():
        if isinstance(value, dict):
            elem = ET.SubElement(parent, key)
            _dict_to_xml(value, elem)
        elif isinstance(value, list):
            for item in value:
                elem = ET.SubElement(parent, key)
                if isinstance(item, dict):
                    _dict_to_xml(item, elem)
                else:
                    elem.text = str(item)
        else:
            elem = ET.SubElement(parent, key)
            elem.text = str(value)


def xml_to_json(xml_data: str) -> str:
    root = ET.fromstring(xml_data)
    data = _xml_to_dict(root)

    if len(data) == 1:
        only_key = list(data.keys())[0]
        data = data[only_key]

    return json.dumps(data, ensure_ascii=False, indent=2)


def _xml_to_dict(element: ET.Element) -> Dict[str, Any]:
    result = {}

    children = list(element)
    if not children:
        return element.text

    child_tags = {}
    for child in children:
        if child.tag in child_tags:
            child_tags[child.tag] += 1
        else:
            child_tags[child.tag] = 1

    for child in children:
        value = _xml_to_dict(child)
        if child_tags[child.tag] > 1:
            if child.tag not in result:
                result[child.tag] = []
            result[child.tag].append(value)
        else:
            result[child.tag] = value

    return result


def _get_cell_value(cell) -> Any:
    value = cell.value
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def excel_to_json(excel_path: str, sheet_name: Optional[str] = None,
                  task_id: Optional[str] = None,
                  chunk_size: int = CHUNK_SIZE) -> str:
    from openpyxl import load_workbook

    if task_id:
        mark_task_started(task_id)
        update_progress(task_id, 5, message='Opening Excel file...')

    file_size = os.path.getsize(excel_path)
    use_read_only = file_size > LARGE_FILE_THRESHOLD

    try:
        wb = load_workbook(excel_path, read_only=use_read_only, data_only=True)
    except Exception as e:
        if task_id:
            mark_task_failed(task_id, str(e))
        raise

    if task_id:
        update_progress(task_id, 10, message='Workbook loaded')

    ws = wb[sheet_name] if sheet_name else wb.active

    all_rows: List[Dict[str, Any]] = []
    headers: List[str] = []
    total_rows = ws.max_row - 1 if ws.max_row else 0
    total_chunks = max(1, (total_rows + chunk_size - 1) // chunk_size) if total_rows > 0 else 1

    if task_id:
        update_progress(task_id, 15, message=f'Reading {total_rows} rows in {total_chunks} chunks',
                        total_chunks=total_chunks)

    row_iter = ws.iter_rows(values_only=False)
    first_row = True
    current_chunk = 0
    rows_in_chunk = 0

    for row in row_iter:
        if first_row:
            headers = [_get_cell_value(cell) for cell in row]
            first_row = False
            continue

        row_data = {}
        for idx, cell in enumerate(row):
            if idx < len(headers):
                row_data[headers[idx]] = _get_cell_value(cell)
        all_rows.append(row_data)
        rows_in_chunk += 1

        if use_read_only and rows_in_chunk >= chunk_size and task_id:
            current_chunk += 1
            progress = 15 + int(75 * current_chunk / total_chunks)
            update_progress(task_id, progress,
                            message=f'Processing chunk {current_chunk}/{total_chunks}',
                            current_chunk=current_chunk)
            rows_in_chunk = 0

    if task_id:
        update_progress(task_id, 95, message='Serializing to JSON...')

    result = json.dumps(all_rows, ensure_ascii=False, indent=2)

    if task_id:
        mark_task_completed(task_id, result=result,
                            message=f'Converted {len(all_rows)} rows successfully')

    wb.close()
    return result


def json_to_excel(json_data: str, excel_path: str,
                  sheet_name: str = 'Sheet1',
                  task_id: Optional[str] = None,
                  chunk_size: int = CHUNK_SIZE) -> None:
    from openpyxl import Workbook

    if task_id:
        mark_task_started(task_id)
        update_progress(task_id, 5, message='Parsing JSON data...')

    data = json.loads(json_data)

    if not isinstance(data, list):
        data = [data]

    if not data:
        if task_id:
            mark_task_failed(task_id, 'Empty JSON data')
        raise ValueError('Empty JSON data')

    total_rows = len(data)
    headers = list(data[0].keys())
    use_write_only = total_rows * len(headers) * 20 > LARGE_FILE_THRESHOLD

    if task_id:
        total_chunks = max(1, (total_rows + chunk_size - 1) // chunk_size)
        update_progress(task_id, 10,
                        message=f'Writing {total_rows} rows in {total_chunks} chunks',
                        total_chunks=total_chunks)

    try:
        wb = Workbook(write_only=use_write_only)
        ws = wb.active
        ws.title = sheet_name

        if task_id:
            update_progress(task_id, 15, message='Writing header row...')

        ws.append(headers)

        current_chunk = 0
        rows_in_chunk = 0
        buffer = []

        for idx, row_data in enumerate(data):
            row_values = [str(row_data.get(h, '')) for h in headers]
            buffer.append(row_values)
            rows_in_chunk += 1

            if rows_in_chunk >= chunk_size:
                for b_row in buffer:
                    ws.append(b_row)
                buffer = []
                rows_in_chunk = 0
                current_chunk += 1
                if task_id:
                    progress = 15 + int(80 * current_chunk / total_chunks)
                    update_progress(task_id, progress,
                                    message=f'Writing chunk {current_chunk}/{total_chunks}',
                                    current_chunk=current_chunk)

        if buffer:
            for b_row in buffer:
                ws.append(b_row)

        if task_id:
            update_progress(task_id, 95, message='Saving Excel file...')

        wb.save(excel_path)
        wb.close()

        if task_id:
            mark_task_completed(task_id, result=excel_path,
                                message=f'Successfully wrote {total_rows} rows to {excel_path}')

    except Exception as e:
        if task_id:
            mark_task_failed(task_id, str(e))
        raise


def convert(source_data: str, source_format: str, target_format: str,
            input_path: Optional[str] = None, output_path: Optional[str] = None,
            task_id: Optional[str] = None) -> Optional[str]:
    source_format = source_format.lower()
    target_format = target_format.lower()

    if source_format == 'xlsx':
        source_format = 'excel'
    if target_format == 'xlsx':
        target_format = 'excel'

    valid_formats = {'csv', 'json', 'xml', 'excel'}
    if source_format not in valid_formats or target_format not in valid_formats:
        raise ValueError(f"Unsupported format. Supported formats: {valid_formats}")

    if source_format == target_format:
        if output_path and source_data:
            write_file(output_path, source_data)
        return source_data

    if source_format == 'excel' and target_format == 'json':
        if not input_path:
            raise ValueError('input_path is required for Excel to JSON conversion')
        return excel_to_json(input_path, task_id=task_id)

    if source_format == 'json' and target_format == 'excel':
        if not output_path:
            raise ValueError('output_path is required for JSON to Excel conversion')
        json_to_excel(source_data, output_path, task_id=task_id)
        return None

    if source_format == 'csv' and target_format == 'json':
        return csv_to_json(source_data)
    elif source_format == 'json' and target_format == 'csv':
        return json_to_csv(source_data)
    elif source_format == 'json' and target_format == 'xml':
        return json_to_xml(source_data)
    elif source_format == 'xml' and target_format == 'json':
        return xml_to_json(source_data)
    elif source_format == 'csv' and target_format == 'xml':
        return json_to_xml(csv_to_json(source_data))
    elif source_format == 'xml' and target_format == 'csv':
        return json_to_csv(xml_to_json(source_data))
    elif source_format == 'csv' and target_format == 'excel':
        if not output_path:
            raise ValueError('output_path is required for CSV to Excel conversion')
        json_data = csv_to_json(source_data)
        json_to_excel(json_data, output_path, task_id=task_id)
        return None
    elif source_format == 'excel' and target_format == 'csv':
        if not input_path:
            raise ValueError('input_path is required for Excel to CSV conversion')
        json_data = excel_to_json(input_path, task_id=task_id)
        return json_to_csv(json_data)
    elif source_format == 'xml' and target_format == 'excel':
        if not output_path:
            raise ValueError('output_path is required for XML to Excel conversion')
        json_data = xml_to_json(source_data)
        json_to_excel(json_data, output_path, task_id=task_id)
        return None
    elif source_format == 'excel' and target_format == 'xml':
        if not input_path:
            raise ValueError('input_path is required for Excel to XML conversion')
        json_data = excel_to_json(input_path, task_id=task_id)
        return json_to_xml(json_data)
    else:
        raise ValueError(f"Cannot convert from {source_format} to {target_format}")


def main():
    import argparse

    if sys.stdout.encoding is None or sys.stdout.encoding.lower() != 'utf-8':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    if sys.stdin.encoding is None or sys.stdin.encoding.lower() not in ('utf-8', 'utf-8-sig'):
        sys.stdin = io.TextIOWrapper(sys.stdin.buffer, encoding='utf-8')

    parser = argparse.ArgumentParser(description='Convert between CSV, JSON, XML, and Excel formats')
    parser.add_argument('source_format', nargs='?',
                        choices=['csv', 'json', 'xml', 'excel', 'xlsx'],
                        help='Source format')
    parser.add_argument('target_format', nargs='?',
                        choices=['csv', 'json', 'xml', 'excel', 'xlsx'],
                        help='Target format')
    parser.add_argument('--input', '-i', help='Input file path (reads from stdin if not provided)')
    parser.add_argument('--output', '-o', help='Output file path (writes to stdout if not provided)')
    parser.add_argument('--task-id', help='Task ID for progress tracking')
    parser.add_argument('--status', metavar='TASK_ID',
                        help='Query task status by task ID (use "all" to list all tasks)')
    parser.add_argument('--list-tasks', action='store_true',
                        help='List all registered tasks')
    parser.add_argument('--sheet-name', help='Sheet name for Excel operations')
    parser.add_argument('--chunk-size', type=int, default=CHUNK_SIZE,
                        help=f'Chunk size for large file processing (default: {CHUNK_SIZE})')

    args = parser.parse_args()

    if args.list_tasks:
        tasks = list_tasks()
        print(json.dumps(tasks, ensure_ascii=False, indent=2, default=str))
        return

    if args.status:
        if args.status.lower() == 'all':
            tasks = list_tasks()
            print(json.dumps(tasks, ensure_ascii=False, indent=2, default=str))
        else:
            status = get_task_status(args.status)
            if status:
                print(json.dumps(status, ensure_ascii=False, indent=2, default=str))
            else:
                print(json.dumps({'error': f'Task {args.status} not found'},
                                 ensure_ascii=False, indent=2))
                sys.exit(1)
        return

    if not args.source_format or not args.target_format:
        parser.error('source_format and target_format are required unless using --status or --list-tasks')

    task_id = args.task_id
    if task_id:
        register_task(task_id)

    source_data = ''
    if args.input:
        if args.source_format in ('excel', 'xlsx'):
            source_data = ''
        else:
            source_data = read_file(args.input)
    else:
        if args.source_format in ('excel', 'xlsx'):
            parser.error('--input is required when source format is excel/xlsx')
        source_data = strip_bom(sys.stdin.read())

    result = convert(
        source_data,
        args.source_format,
        args.target_format,
        input_path=args.input,
        output_path=args.output,
        task_id=task_id,
    )

    if result is not None:
        if args.output and args.target_format not in ('excel', 'xlsx'):
            write_file(args.output, result)
        elif args.target_format not in ('excel', 'xlsx'):
            print(result)

    if task_id:
        status = get_task_status(task_id)
        if status:
            print(f'\n[Task {task_id}] {status["status"]} - {status["message"]}')
        sys.exit(0 if status and status['status'] == 'completed' else 1)


if __name__ == '__main__':
    main()
